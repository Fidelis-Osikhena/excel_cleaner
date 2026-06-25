from __future__ import annotations

import math

from openpyxl import Workbook

from processor import (
    clean_orientation,
    ensure_column_after,
    fill_column_green,
    get_headers,
    normalize_feature_type,
    orientation_to_degrees,
    parse_yes,
)


ENBRIDGE_COLUMNS = [
    "Type of Object",
    "Feature ID",
    "Odometer (m)",
    "NPS",
    "Tool Technology",
    "Feature Type",
    "Feature Identification",
    "Depth (%)",
    "Depth (mm)",
    "Length (mm)",
    "Width (mm)",
    "Clock Position",
    "Maximum Strain",
    "Cluster ID",
    "Surface Location",
    "Weld Interaction (Y/N)",
    "Joint Number",
    "Long Seam Position Start",
    "Long Seam Position End",
    "Wall Thickness (mm)",
    "Tool Passage (hh:mm:sec)",
    "Tool Velocity (m/s)",
    "Degraded Data? (Y/N)",
    "Feature Morphology Classification",
    "Manual Anomaly Assessment (Y/N)",
    "Depth Tolerance (±%)",
    "Length Tolerance (±mm)",
    "Width Tolerance (±mm)",
    "Comments",
    "Latitude (deg)",
    "Longitude (deg)",
    "Elevation (m)",
    "Pipe Grade (MPa)",
    "Seam Type",
    "CVN (J)",
    "Pipeline MOP (psi)",
    "Distance to U/S GW (m)",
    "Distance to long seam weld (mm)",
    "Effective Area failure pressure (psi)",
    "Mod B31G failure pressure (psi)",
    "CorLAS failure pressure (psi)",
    "CorLAS RPR",
    "Mod B31G RPR",
    "Effective Area RPR",
    "CorLAS FPR",
    "Mod B31G FPR",
    "Effective Area FPR",
    "Depth growth rate (mm/y)",
    "Length growth rate (mm/y)",
    "Width growth rate (mm/y)",
    "Matching Comment",
    "Coating Disbondment (%)",
    "Effective Depth (mm)",
    "Effective Length (mm)",
    "25mm from GW?",
    "25mm from SW?",
]


def process_enbridge_canada_workbook(
    wb: Workbook,
    pipe_diameter: float | None = None,
) -> None:
    ws = wb.active

    headers, output_map = _ensure_enbridge_columns(ws)
    _process_enbridge_rows(ws, headers, output_map, pipe_diameter)


def _ensure_enbridge_columns(ws) -> tuple[dict[str, int], dict[str, str]]:
    headers = get_headers(ws)
    output_map: dict[str, str] = {}

    anchor = "Feature Type"

    for desired_name in ENBRIDGE_COLUMNS:
        actual_name = get_available_output_column_name(headers, desired_name)

        ensure_column_after(ws, anchor, actual_name, headers)
        headers = get_headers(ws)

        output_map[desired_name] = actual_name
        anchor = actual_name

    return headers, output_map


def _process_enbridge_rows(ws,headers: dict[str, int],output_map: dict[str, str],pipe_diameter: float | None,) -> None:
    feature_col = headers.get("feature type")
    if feature_col is None:
        return

    for row in range(2, ws.max_row + 1):
        source_feature_value = ws.cell(row=row, column=feature_col).value
        feature = normalize_feature_type(source_feature_value)

        values = map_enbridge_feature_row(ws, row, headers, feature, pipe_diameter)

        for desired_header, value in values.items():
            actual_header = output_map.get(desired_header, desired_header)
            col = headers.get(actual_header.strip().lower())

            if col:
                ws.cell(row=row, column=col).value = value

    for desired_header in ENBRIDGE_COLUMNS:
        actual_header = output_map.get(desired_header, desired_header)
        col = headers.get(actual_header.strip().lower())

        if col:
            fill_column_green(ws, col)


def map_enbridge_feature_row(
    ws,
    row: int,
    headers: dict[str, int],
    feature: str,
    pipe_diameter: float | None,
) -> dict[str, object]:
    # This is where we add feature-specific mapping rules.
    result: dict[str, object] = {}

    result["Type of Object"] = map_enbridge_type_of_object(feature)
    result["Feature Type"] = map_enbridge_feature_type(feature)
    result["Feature Identification"] = map_enbridge_feature_identification(feature)

    return result


def map_enbridge_type_of_object(feature: str) -> str:
    feature_type_object_map = {
        "bend": "F",
        "cp attachment": "F",
        "crack-like": "F",
        "deformation": "F",
        "deformation w/ metal loss": "F",
        "half sole repair": "F",
        "manufacturing anomaly": "F",
        "metal loss": "F",
        "metal loss cluster": "F",
        "metal loss manufacturing": "F",
        "metal loss manufacturing cluster": "F",
        "metal object - close": "F",
        "metal object - touching": "F",
        "patch repair": "F",
        "pipe support - rectangular": "F",
        "puddle weld repair": "F",
        "repair marker begin": "F",
        "repair marker end": "F",
        "stopple": "F",
        "swa": "F",
        "swa cluster": "F",
        "swf-a": "F",
        "swf-b": "F",
        "tap": "F",
        "tee": "F",
        "weld anomaly": "F",
        "casing begin": "F",
        "casing end": "F",
        "flange": "F",
        "marker band begin": "F",
        "marker band end": "F",
        "pipe support - circumferential": "F",
        "recoat begin": "F",
        "recoat end": "F",
        "sleeve begin": "F",
        "sleeve end": "F",
        "valve": "F",
        "agm": "M",
        "girth weld": "S",
    }

    return feature_type_object_map.get(feature, "")


def map_enbridge_feature_type(feature: str) -> str:
    feature_type_map = {
        "bend": "Bend",
        "cp attachment": "CP connection",
        "crack-like": "Anomaly",
        "deformation": "Anomaly",
        "deformation w/ metal loss": "Anomaly",
        "half sole repair": "Repair",
        "manufacturing anomaly": "Anomaly",
        "metal loss": "Anomaly",
        "metal loss cluster": "Anomaly",
        "metal loss manufacturing": "Anomaly",
        "metal loss manufacturing cluster": "Anomaly",
        "metal object - close": "Additional metal/material",
        "metal object - touching": "Additional metal/material",
        "patch repair": "Repair",
        "pipe support - rectangular": "External support",
        "puddle weld repair": "Repair",
        "repair marker begin": "Repair",
        "repair marker end": "Repair",
        "stopple": "Stopper",
        "swa": "Anomaly",
        "swa cluster": "Anomaly",
        "swf-a": "Anomaly",
        "swf-b": "Anomaly",
        "tap": "Tap",
        "tee": "Tee",
        "weld anomaly": "Anomaly",
        "agm": "Above Ground Marker",
        "casing begin": "Casing begin",
        "casing end": "Casing end",
        "flange": "Pipeline fixture",
        "girth weld": "Girth Weld",
        "marker band begin": "Repair",
        "marker band end": "Repair",
        "pipe support - circumferential": "External support",
        "recoat begin": "Repair",
        "recoat end": "Repair",
        "sleeve begin": "Repair",
        "sleeve end": "Repair",
        "valve": "Valve",
    }

    return feature_type_map.get(feature, "")


def map_enbridge_feature_identification(feature: str) -> str:
    feature_identification_map = {
        "metal loss": "Metal loss",
        "metal loss cluster": "Metal loss cluster",
        "metal loss manufacturing": "Pipe mill anomaly",
        "metal loss manufacturing cluster": "Pipe mill anomaly cluster",
        "manufacturing anomaly": "Pipe mill anomaly",
        "deformation": "Dent",
        "deformation w/ metal loss": "Dent with metal loss",
        "metal object - close": "Close Metal Object",
        "metal object - touching": "Touching Metal Object",
        "patch repair": "Other",
        "puddle weld repair": "Puddle Weld Repair",
        "repair marker begin": "Other begin",
        "repair marker end": "Other end",
        "swa": "Longitudinal weld anomaly",
        "swa cluster": "Longitudinal weld anomaly",
        "swf-a": "Longitudinal weld crack",
        "swf-b": "Longitudinal weld crack",
        "weld anomaly": "Girth weld anomaly",
        "flange": "Flange",
        "girth weld": "Not identifiable seam",
        "marker band begin": "Composite sleeve begin",
        "marker band end": "Composite sleeve end",
        "recoat begin": "Other",
        "recoat end": "Other",
        "sleeve begin": "Welded sleeve begin",
        "sleeve end": "Welded sleeve end",
    }

    return feature_identification_map.get(feature, "")

def map_enbridge_feature_row(
    ws,
    row: int,
    headers: dict[str, int],
    feature: str,
    pipe_diameter: float | None,
) -> dict[str, object]:
    result: dict[str, object] = {}

    depth_percent = get_depth_percent(get_source(ws, row, headers, ["Depth (%)", "Depth %", "Depth"]))
    wall_thickness = to_float(get_source(ws, row, headers, ["Wall Thickness", "Wall Thickness (mm)"]))
    length_value = get_source(ws, row, headers, ["Length (mm)", "Length (in)", "Length"])
    width_value = get_source(ws, row, headers, ["Width (mm)", "Width (in)", "Width"])
    diameter = pipe_diameter

    result["Type of Object"] = map_enbridge_type_of_object(feature)
    result["Feature Type"] = map_enbridge_feature_type(feature)
    result["Feature Identification"] = map_enbridge_feature_identification(feature)

    result["Feature ID"] = get_source(ws, row, headers, ["Feature ID", "Feature Id"])
    result["Odometer (m)"] = get_source(ws, row, headers, ["Odometer Main (ft)", "Odometer Main (m)", "Odometer"])
    result["NPS"] = get_nps(pipe_diameter)
    result["Tool Technology"] = map_tool_technology(get_source(ws, row, headers, ["Sensor Type"]))
    result["Clock Position"] = get_source(ws, row, headers, ["Orientation Main", "Orientation Final", "Orientation Center"])
    result["Surface Location"] = map_surface_location(get_source(ws, row, headers, ["IsExternal", "Is External"]))
    result["Joint Number"] = make_joint_number(get_source(ws, row, headers, ["Weld Id", "Weld ID"]))
    result["Long Seam Position Start"] = get_source(ws, row, headers, ["Long Seam Orientation"])
    result["Long Seam Position End"] = get_source(ws, row, headers, ["Long Seam Orientation"])
    result["Wall Thickness (mm)"] = wall_thickness if wall_thickness is not None else ""
    result["Tool Passage (hh:mm:sec)"] = get_source(ws, row, headers, ["Tool Passage Time (UTC time)", "Tool Passage Time", "Tool Passage"])
    result["Tool Velocity (m/s)"] = get_source(ws, row, headers, ["Speed", "Tool Speed"])
    result["Latitude (deg)"] = get_source(ws, row, headers, ["Latitude", "Lat"])
    result["Longitude (deg)"] = get_source(ws, row, headers, ["Longitude", "Long"])
    result["Elevation (m)"] = get_source(ws, row, headers, ["Height", "Height (m)", "Elevation"])
    result["Pipe Grade (MPa)"] = divide_or_blank(get_source(ws, row, headers, ["SMYS", "SMYS (kPa)"]), 1000)
    result["Seam Type"] = get_source(ws, row, headers, ["Seam Type"])
    result["Pipeline MOP (psi)"] = multiply_or_blank(get_source(ws, row, headers, ["MOP", "MOP (kPa)"]), 0.145038)
    result["Distance to U/S GW (m)"] = get_source(ws, row, headers, ["US Weld Δ main", "US Weld Δ main (m)", "US Weld ? main"])
    result["Distance to long seam weld (mm)"] = ""

    result["Effective Area failure pressure (psi)"] = multiply_or_blank(get_source(ws, row, headers, ["Rstreng", "RSTRENG"]), 0.145038)
    result["Mod B31G failure pressure (psi)"] = multiply_or_blank(get_source(ws, row, headers, ["MB31G", "MB31G (kPa)"]), 0.145038)
    result["CorLAS failure pressure (psi)"] = ""

    result["Mod B31G RPR"] = calc_pressure_ratio(
        get_source(ws, row, headers, ["MB31G", "MB31G (kPa)"]),
        get_source(ws, row, headers, ["SMYS", "SMYS (kPa)"]),
        wall_thickness,
        diameter,
    )
    result["Effective Area RPR"] = calc_pressure_ratio(
        get_source(ws, row, headers, ["Rstreng", "RSTRENG"]),
        get_source(ws, row, headers, ["SMYS", "SMYS (kPa)"]),
        wall_thickness,
        diameter,
    )
    result["CorLAS RPR"] = ""

    mop_raw = get_source(ws, row, headers, ["MOP", "MOP (kPa)"])
    result["Mod B31G FPR"] = calc_fpr_enbridge(get_source(ws, row, headers, ["MB31G", "MB31G (kPa)"]), mop_raw)
    result["Effective Area FPR"] = calc_fpr_enbridge(get_source(ws, row, headers, ["Rstreng", "RSTRENG"]), mop_raw)
    result["CorLAS FPR"] = ""

    result["Effective Depth (mm)"] = calc_effective_depth_mm(
        get_source(ws, row, headers, ["Effective Depth (%)", "Effective Depth"]),
        wall_thickness,
    )
    result["Effective Length (mm)"] = get_source(ws, row, headers, ["Effective Length", "Effective Length (mm)"])

    # Feature-specific dimensions/depth/comments
    apply_enbridge_feature_specifics(
        result=result,
        ws=ws,
        row=row,
        headers=headers,
        feature=feature,
        depth_percent=depth_percent,
        wall_thickness=wall_thickness,
        pipe_diameter=pipe_diameter,
        length_value=length_value,
        width_value=width_value,
    )

    result["25mm from GW?"] = get_25mm_from_gw(ws, row, headers)
    result["25mm from SW?"] = get_25mm_from_sw(ws, row, headers, pipe_diameter)

    return result

def apply_enbridge_feature_specifics(
    *,
    result: dict[str, object],
    ws,
    row: int,
    headers: dict[str, int],
    feature: str,
    depth_percent: float | None,
    wall_thickness: float | None,
    pipe_diameter: float | None,
    length_value,
    width_value,
) -> None:
    # Depth (%) is output as whole-number percent where applicable
    if depth_percent is not None:
        result["Depth (%)"] = round(depth_percent)

    # Depth in mm
    if feature in {
        "metal loss", "metal loss cluster",
        "metal loss manufacturing", "metal loss manufacturing cluster",
        "swa", "swa cluster", "swf-a", "swf-b", "crack-like",
    }:
        result["Depth (mm)"] = round(depth_percent / 100 * wall_thickness, 2) if depth_percent is not None and wall_thickness else ""

    elif feature in {"deformation", "deformation - ovality", "deformation w/ metal loss"}:
        result["Depth (mm)"] = round(depth_percent / 100 * pipe_diameter, 2) if depth_percent is not None and pipe_diameter else ""

    # Length / Width
    if feature == "cp attachment":
        result["Length (mm)"] = round_to_13_or_25(length_value)
        result["Width (mm)"] = round_to_13_or_25(width_value)

    elif feature in {
        "crack-like", "deformation", "deformation - ovality", "deformation w/ metal loss",
        "manufacturing anomaly", "metal loss", "metal loss cluster",
        "metal loss manufacturing", "metal loss manufacturing cluster",
        "metal object - close", "metal object - touching",
        "patch repair", "pipe support - rectangular", "puddle weld repair",
        "swa", "swa cluster", "swf-a", "swf-b", "weld anomaly",
    }:
        result["Length (mm)"] = length_value
        result["Width (mm)"] = width_value

    # Comments
    source_comment = safe_str(get_source(ws, row, headers, ["Comments", "Comment Working", "Comment"]))

    if feature == "bend":
        result["Comments"] = enbridge_bend_comment(ws, row, headers)

    elif feature in {"half sole repair", "patch repair", "pipe support - rectangular", "repair marker begin", "repair marker end", "pipe support - circumferential", "recoat begin", "recoat end"}:
        result["Comments"] = safe_str(get_source(ws, row, headers, ["Feature Type"]))

    elif feature == "stopple":
        result["Comments"] = f"Stopple size: {round_numeric(length_value)} mm"

    elif feature == "tap":
        result["Comments"] = f"Tap diameter: {round_numeric(length_value)} mm"

    elif feature == "tee":
        base = f"Tee size: {round_numeric(length_value)} mm"
        result["Comments"] = f"{base}. {source_comment}" if source_comment else base

    elif feature in {"agm", "flange", "valve"}:
        result["Comments"] = source_comment


def get_source(ws, row: int, headers: dict[str, int], names: list[str]):
    for name in names:
        col = headers.get(name.strip().lower())
        if col:
            value = ws.cell(row=row, column=col).value
            return "" if value is None else value
    return ""


def safe_str(value: object) -> str:
    return "" if value is None else str(value).strip()


def to_float(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def get_nps(pipe_diameter: float | None) -> str:
    if pipe_diameter is None:
        return ""
    return str(int(pipe_diameter))


def map_tool_technology(sensor_value: object) -> str:
    sensor = safe_str(sensor_value).lower()
    if sensor == "axial":
        return "MFL"
    if sensor == "circumferential":
        return "CMFL"
    if sensor == "idd-sm":
        return "IDD-SM"
    if sensor == "geometry":
        return "Geometry"
    return ""


def map_surface_location(value: object) -> str:
    text = safe_str(value).lower()
    if text == "yes":
        return "EXT"
    if text == "no":
        return "INT"
    return ""


def make_joint_number(weld_id: object) -> str:
    weld = safe_str(weld_id)
    return f"ILI.GW.{weld}" if weld else ""


def multiply_or_blank(value: object, factor: float):
    num = to_float(value)
    return "" if num is None else num * factor


def divide_or_blank(value: object, divisor: float):
    num = to_float(value)
    return "" if num is None or divisor == 0 else num / divisor


def calc_pressure_ratio(pressure, smys, wall_thickness, diameter):
    pressure_num = to_float(pressure)
    smys_num = to_float(smys)
    wt_num = to_float(wall_thickness)
    dia_num = to_float(diameter)

    if None in (pressure_num, smys_num, wt_num, dia_num):
        return ""
    if smys_num == 0 or wt_num == 0 or dia_num == 0:
        return ""

    return round(pressure_num / ((2 * wt_num * smys_num) / dia_num), 2)


def calc_fpr_enbridge(pressure, mop):
    pressure_num = to_float(pressure)
    mop_num = to_float(mop)

    if pressure_num is None or mop_num in (None, 0):
        return ""

    return round(pressure_num / mop_num, 2)


def calc_effective_depth_mm(effective_depth_percent, wall_thickness):
    depth = to_float(effective_depth_percent)
    wt = to_float(wall_thickness)

    if depth is None or wt is None:
        return ""

    if 0 < depth < 1:
        depth = depth * 100

    return round((depth / 100) * wt, 2)


def round_to_13_or_25(value):
    num = to_float(value)
    if num is None:
        return ""
    return 13 if abs(num - 13) <= abs(num - 25) else 25


def round_numeric(value):
    num = to_float(value)
    return "" if num is None else round(num)


def enbridge_bend_comment(ws, row: int, headers: dict[str, int]) -> str:
    radius = format_one_decimal(get_source(ws, row, headers, ["Bend Radius (xD)"]))
    angle = format_one_decimal(get_source(ws, row, headers, ["Bend Angle"]))
    direction = bend_orientation_to_direction(get_source(ws, row, headers, ["Bend Orientation"]))

    base = f"{radius}D - {angle}\u00B0 - {direction}".strip()
    seam_type = safe_str(get_source(ws, row, headers, ["Seam Type"])).lower()
    comment = safe_str(get_source(ws, row, headers, ["Comments", "Comment"]))

    if seam_type == "forged elbow" and comment:
        return f"{base}, {comment}"

    return base


def format_one_decimal(value: object) -> str:
    num = to_float(value)
    return "" if num is None else f"{num:.1f}"


def bend_orientation_to_direction(value: object) -> str:
    cleaned = clean_orientation(value)
    if not cleaned:
        return ""

    hour, minute = map(int, cleaned.split(":"))
    if hour == 12:
        hour = 0

    total_minutes = hour * 60 + minute

    if total_minutes >= 630 or total_minutes <= 89:
        return "up"
    if total_minutes <= 269:
        return "right"
    if total_minutes <= 449:
        return "down"
    return "left"

def get_depth_percent(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip().replace("%", "")
    if not text:
        return None

    try:
        num = float(text)
    except (TypeError, ValueError):
        return None

    # Excel may store 1.5% as 0.015
    if 0 < num < 1:
        return num * 100

    return num

def get_25mm_from_gw(ws, row: int, headers: dict[str, int]) -> str:
    """
    Returns text like:
    11mm away from Girth Weld

    Rule:
    If the minimum weld edge distance is <= 0.025 m,
    OR US Weld Δ from right edge < US Weld Δ from left edge,
    then return the minimum distance converted to mm.
    """

    us_left = to_float(get_source(ws, row, headers, [
        "US Weld Δ from left edge",
        "US Weld Δ from left edge (m)",
        "US Weld ? from left edge",
    ]))

    ds_left = to_float(get_source(ws, row, headers, [
        "DS Weld Δ from left edge",
        "DS Weld Δ from left edge (m)",
        "DS Weld ? from left edge",
    ]))

    us_right = to_float(get_source(ws, row, headers, [
        "US Weld Δ from right edge",
        "US Weld Δ from right edge (m)",
        "US Weld ? from right edge",
    ]))

    ds_right = to_float(get_source(ws, row, headers, [
        "DS Weld Δ from right edge",
        "DS Weld Δ from right edge (m)",
        "DS Weld ? from right edge",
    ]))

    values = [v for v in [us_left, ds_left, us_right, ds_right] if v is not None]

    if not values:
        return ""

    min_m = min(values)

    is_close = min_m <= 0.025
    overlaps_us_weld = (
        us_right is not None
        and us_left is not None
        and us_right < us_left
    )

    if is_close or overlaps_us_weld:
        return f"{round(min_m * 1000)}mm away from Girth Weld"

    return ""

def get_25mm_from_sw(ws, row: int, headers: dict[str, int], pipe_diameter: float | None) -> str:
    """
    Returns text like:
    6mm away from Seam Weld

    Rule:
    Find the shortest circumferential distance from either top-left or
    top-right orientation to the long seam orientation.
    If the shortest distance is <= 25 mm, return that distance.
    """

    if pipe_diameter is None:
        return ""

    diameter_mm = pipe_diameter_to_mm(pipe_diameter)
    if diameter_mm is None:
        return ""

    long_seam_deg = get_orientation_degrees_from_sources(
        ws,
        row,
        headers,
        [
            "Long Seam Orientation in Degrees",
            "Long Seam Orientation (Degree)",
            "Long Seam Orientation",
        ],
    )

    top_left_deg = get_orientation_degrees_from_sources(
        ws,
        row,
        headers,
        [
            "Top Left Orientation in Degrees",
            "Top Left Orientation (Degree)",
            "Top Left Orientation",
        ],
    )

    top_right_deg = get_orientation_degrees_from_sources(
        ws,
        row,
        headers,
        [
            "Top Right Orientation in Degrees",
            "Top Right Orientation (Degree)",
            "Top Right Orientation",
        ],
    )

    if long_seam_deg is None:
        return ""

    distances = []

    if top_left_deg is not None:
        distances.append(circumferential_distance_mm(top_left_deg, long_seam_deg, diameter_mm))

    if top_right_deg is not None:
        distances.append(circumferential_distance_mm(top_right_deg, long_seam_deg, diameter_mm))

    if not distances:
        return ""

    min_distance_mm = min(distances)
    rounded_distance = round(min_distance_mm)

    if rounded_distance <= 25:
        return f"{rounded_distance}mm away from Seam Weld"

    return ""

def pipe_diameter_to_mm(pipe_diameter: float | None) -> float | None:
    """
    Converts pipe diameter from inches to mm.
    The GUI pipe diameter dropdown is in inches.
    """

    if pipe_diameter is None:
        return None

    return pipe_diameter * 25.4

def get_orientation_degrees_from_sources(
    ws,
    row: int,
    headers: dict[str, int],
    names: list[str],
) -> float | None:
    """
    Reads an orientation from one of several possible columns.

    Handles:
    - numeric degrees
    - clock values like 3:00
    """

    value = get_source(ws, row, headers, names)

    if value in ("", None):
        return None

    # Already numeric degrees
    numeric = to_float(value)
    if numeric is not None:
        return numeric

    # Clock format
    cleaned = clean_orientation(value)
    if not cleaned:
        return None

    try:
        hour, minute = map(int, cleaned.split(":"))
    except ValueError:
        return None

    if hour == 12:
        hour = 0

    return (hour * 30) + (minute * 0.5)

def circumferential_distance_mm(
    orientation_deg: float,
    seam_deg: float,
    diameter_mm: float,
) -> float:
    """
    Formula:
    min angular difference * pi * diameter / 360
    """

    diff = abs(orientation_deg - seam_deg)

    if diff > 180:
        diff = 360 - diff

    return diff * math.pi * diameter_mm / 360

def get_available_enbridge_column_name(
    headers: dict[str, int],
    desired_name: str,
) -> str:
    """
    If desired_name already exists,
    create:
        Name (New)
        Name (New 2)
        Name (New 3)
        ...
    """

    if desired_name.strip().lower() not in headers:
        return desired_name

    candidate = f"{desired_name} (New)"
    if candidate.strip().lower() not in headers:
        return candidate

    counter = 2

    while True:
        candidate = f"{desired_name} (New {counter})"

        if candidate.strip().lower() not in headers:
            return candidate

        counter += 1

def get_available_output_column_name(
    headers: dict[str, int],
    desired_name: str,
) -> str:
    desired_key = desired_name.strip().lower()

    if desired_key not in headers:
        return desired_name

    candidate = f"{desired_name} (New)"
    if candidate.strip().lower() not in headers:
        return candidate

    counter = 2
    while True:
        candidate = f"{desired_name} (New {counter})"
        if candidate.strip().lower() not in headers:
            return candidate
        counter += 1