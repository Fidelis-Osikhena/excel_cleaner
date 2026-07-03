from __future__ import annotations

import os
import csv
from openpyxl import load_workbook


JOINT_OUTPUT_COLUMNS = [
    "Upstream Girth Weld Number",
    "Previous US Girth Weld Number",
    "Odometer",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Height",
    "Wall Thickness",
    "Grade",
    "Diameter",
    "Seam Type",
    "Seam Position",
    "Comments",
    "MOP",
    "DPP",
    "Tool Speed",
    "Detectable Length",
]


def generate_dras_files(
    excel_path: str,
    output_folder: str,
    pipe_diameter: float,
) -> None:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    os.makedirs(output_folder, exist_ok=True)

    generate_joint_txt(ws, output_folder, pipe_diameter)
    generate_cluster_txt(ws, output_folder, pipe_diameter)
    generate_callbox_txt(ws, output_folder, pipe_diameter)
    generate_crack_anomalies_txt(ws, output_folder, pipe_diameter)
    generate_facilities_txt(ws, output_folder)
    generate_other_anomalies_txt(ws, output_folder)
    generate_bend_strain_txt(output_folder)
    generate_inline_inspection_txt(output_folder)


def generate_joint_txt(ws, output_folder: str, pipe_diameter: float) -> None:
    """
    Generates Joint.txt.

    Only imports rows where Joint = 1.
    Output is tab-delimited.
    """

    headers = get_headers(ws)

    output_path = os.path.join(output_folder, "Joint.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=JOINT_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )

        writer.writeheader()

        for row in range(2, ws.max_row + 1):
            joint_value = get_value(ws, row, headers, ["Joint.txt"])

            if not is_joint_row(joint_value):
                continue

            output_row = {
                "Upstream Girth Weld Number": get_value(ws, row, headers, ["Weld Id", "Weld ID"]),
                "Previous US Girth Weld Number": "",
                "Odometer": get_value(ws, row, headers, ["Odometer Main (m)", "Odometer Main (ft)"]),
                "X_Coord": get_value(ws, row, headers, ["Easting (m)", "Easting (ft)"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (m)", "Northing (ft)"]),
                "Lat": get_value(ws, row, headers, ["Latitude"]),
                "Long": get_value(ws, row, headers, ["Longitude"]),
                "Height": get_value(ws, row, headers, ["Height (m)", "Height (ft)"]),
                "Wall Thickness": get_value(ws, row, headers, ["Wall Thickness (mm)", "Wall Thickness (in)"]),
                "Grade": get_value(ws, row, headers, ["SMYS (kPa)", "SMYS (psi)"]),
                "Diameter": pipe_diameter,
                "Seam Type": get_dras_seam_type(ws, row, headers),
                "Seam Position": get_dras_seam_position(ws, row, headers),
                "Comments": get_dras_joint_comment(ws, row, headers),
                "MOP": get_mop_value(ws, row, headers),
                "DPP": get_value(ws, row, headers, ["DPP"]),
                "Tool Speed": get_value(ws, row, headers, ["Speed (m/s)", "Speed (ft/s)"]),
                "Detectable Length": get_detectable_length(ws, row, headers),
            }

            writer.writerow(clean_output_row(output_row))


CLUSTER_OUTPUT_COLUMNS = [
    "ID",
    "Number of Boxes in Cluster",
    "Odometer",
    "Azimuth",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Height",
    "Length",
    "Width",
    "Peak Depth",
    "Effective Length",
    "Effective Depth",
    "Failure Pressure",
    "FPR",
    "FPRTC",
    "RPR",
    "Due Date",
    "Status",
    "Description",
    "Surface",
    "Metal Loss Type",
    "Growth Rate",
    "ERF",
    "Depth Tolerance - @ 80% Conf.",
    "Depth Tolerance - StdDev",
    "Length Tolerance - @ 80% Conf.",
    "Length Tolerance - StdDev",
    "Width Tolerance - @ 80% Conf.",
    "Width Tolerance - StdDev",
]


def generate_cluster_txt(ws, output_folder: str, pipe_diameter: float) -> None:
    headers = get_headers(ws)
    output_path = os.path.join(output_folder, "Cluster.txt")

    cluster_counts = {}

    for r in range(2, ws.max_row + 1):
        cluster_value = get_value(ws, r, headers, ["Cluster"])

        if cluster_value in ("", None):
            continue

        cluster_key = str(cluster_value).strip()
        cluster_counts[cluster_key] = cluster_counts.get(cluster_key, 0) + 1

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=CLUSTER_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()

        for row in range(2, ws.max_row + 1):
            if not is_flagged_row(ws, row, headers, ["Cluster.txt"]):
                continue

            id_value = get_value(ws, row, headers, ["ID", "Id"])
            feature_id = get_value(ws, row, headers, ["Feature ID", "Feature Id"])

            id_key = "" if id_value in ("", None) else str(id_value).strip()
            num_boxes = cluster_counts.get(id_key, "")

            maop = get_value(ws, row, headers, ["MAOP (kPa)", "MAOP (psi)", "MAOP"])
            mop = get_value(ws, row, headers, ["MOP (kPa)", "MOP (psi)", "MOP"])
            smys = get_value(ws, row, headers, ["SMYS (kPa)", "SMYS (psi)", "SMYS"])
            mb31g = get_value(ws, row, headers, ["MB31G (kPa)", "MB31G (psi)", "MB31G"])
            wall_thickness = get_value(ws, row, headers, ["Wall Thickness (mm)", "Wall Thickness (in)", "Wall Thickness", "WT"])

            output_row = {
                "ID": feature_id,
                "Number of Boxes in Cluster": num_boxes,
                "Odometer": get_value(ws, row, headers, ["Odometer Main (m)", "Odometer Main (ft)", "Odometer Main", "Odometer"]),
                "Azimuth": get_callbox_azimuth(ws, row, headers),
                "X_Coord": get_value(ws, row, headers, ["Easting (m)", "Easting (ft)", "Easting", "X_Coord", "X Coord"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (m)", "Northing (ft)", "Northing", "Y_Coord", "Y Coord"]),
                "Lat": get_value(ws, row, headers, ["Latitude", "Lat"]),
                "Long": get_value(ws, row, headers, ["Longitude", "Long"]),
                "Height": get_value(ws, row, headers, ["Height (m)", "Height (ft)", "Height"]),
                "Length": get_value(ws, row, headers, ["Length (mm)", "Length (in)", "Length"]),
                "Width": get_value(ws, row, headers, ["Width (mm)", "Width(mm)", "Width (in)", "Width"]),
                "Peak Depth": get_value(ws, row, headers, ["Depth (%)", "Peak Depth"]),
                "Effective Length": get_value(ws, row, headers, ["Effective Length (mm)", "Effective Length (in)", "Effective Length"]),
                "Effective Depth": format_decimal(get_value(ws, row, headers, ["Effective Depth (%)", "Effective Depth"]), 2),
                "Failure Pressure": mb31g,
                "FPR": calc_fpr_choose_pressure(mb31g, maop, mop),
                "FPRTC": get_existing_export_value(ws, row, headers, "FPRTC"),
                "RPR": calc_rpr(mb31g, smys, wall_thickness, pipe_diameter),
                "Due Date": "",
                "Status": get_existing_export_value(ws, row, headers, "Status"),
                "Description": get_dras_description(ws, row, headers),
                "Surface": get_surface(ws, row, headers),
                "Metal Loss Type": get_existing_export_value(ws, row, headers, "Metal Loss Type"),
                "Growth Rate": "",
                "ERF": calc_erf(maop, mop, mb31g),
                "Depth Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]),
                "Depth Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]), 2),
                ),
                "Length Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]),
                "Length Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]), 2),
                ),
                "Width Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]),
                "Width Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]), 2),
                ),
            }

            writer.writerow(clean_output_row(output_row))

CALLBOX_OUTPUT_COLUMNS = [
    "ID",
    "Cluster ID",
    "Odometer",
    "Azimuth",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Height",
    "Length",
    "Width",
    "Peak Depth",
    "Failure Pressure",
    "FPR",
    "FPRTC",
    "RPR",
    "Manual Analysis Flag",
    "Surface",
    "Metal Loss Type",
    "Growth Rate",
    "Discovery Date",
    "ERF",
    "Depth Tolerance - @ 80% Conf.",
    "Depth Tolerance - StdDev",
    "Length Tolerance - @ 80% Conf.",
    "Length Tolerance - StdDev",
    "Width Tolerance - @ 80% Conf.",
    "Width Tolerance - StdDev",
]


def generate_callbox_txt(ws, output_folder: str, pipe_diameter: float) -> None:
    headers = get_headers(ws)
    output_path = os.path.join(output_folder, "Callbox.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=CALLBOX_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()
        id_lookup = {}

        for r in range(2, ws.max_row + 1):
            row_id = get_value(ws, r, headers, ["ID"])
            feature_id = get_value(ws, r, headers, ["Feature ID"])

            if row_id not in ("", None):
                id_lookup[str(row_id).strip()] = feature_id


        for row in range(2, ws.max_row + 1):
            if not is_flagged_row(ws, row, headers, ["Callbox.txt", "Call Box"]):
                continue
            feature_id = get_value(ws, row, headers, ["Feature ID", "Feature Id"])
            if feature_id in ("", None):
                continue

            maop = get_value(ws, row, headers, ["MAOP (kPa)", "MAOP (psi)"])
            mop = get_value(ws, row, headers, ["MOP (kPa)", "MOP (psi)"])
            smys = get_value(ws, row, headers, ["SMYS (kPa)", "SMYS (psi)"])
            mb31g = get_value(ws, row, headers, ["MB31G (kpa)", "MB31G (psi)"])
            wall_thickness = get_value(ws, row, headers, ["Wall Thickness (mm)", "Wall Thickness (in)"])

            output_row = {
                "ID": feature_id,
                "Cluster ID": get_cluster_feature_id(ws, row, headers,id_lookup),
                "Odometer": get_value(ws, row, headers, ["Odometer Main (ft)", "Odometer Main (m)"]),
                "Azimuth": get_callbox_azimuth(ws, row, headers),
                "X_Coord": get_value(ws, row, headers, ["Easting (ft)", "Easting (m)"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (ft)", "Northing (m)"]),
                "Lat": get_value(ws, row, headers, ["Latitude"]),
                "Long": get_value(ws, row, headers, ["Longitude"]),
                "Height": get_value(ws, row, headers, ["Height (ft)", "Height (m)"]),
                "Length": get_value(ws, row, headers, ["Length (mm)", "Length (in)"]),
                "Width": get_value(ws, row, headers, ["Width (mm)", "Width (in)"]),
                "Peak Depth": get_value(ws, row, headers, ["Depth (%)", "Peak Depth"]),
                "Failure Pressure": mb31g,
                "FPR": calc_fpr_choose_pressure(mb31g, maop, mop),
                "FPRTC": get_existing_export_value(ws, row, headers, "FPRTC"),
                "RPR": calc_rpr(mb31g, smys, wall_thickness, pipe_diameter),
                "Manual Analysis Flag": get_existing_export_value(ws, row, headers, "Manual Analysis Flag"),
                "Surface": get_surface(ws, row, headers),
                "Metal Loss Type": get_existing_export_value(ws, row, headers, "Metal Loss Type"),
                "Growth Rate": "",
                "Discovery Date": "",
                "ERF": calc_erf(maop, mop, mb31g),
                "Depth Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]),
                "Depth Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]), 2),
                ),

                "Length Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]),
                "Length Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]), 2),
                ),

                "Width Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]),
                "Width Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]), 2),
                ),
            }

            writer.writerow(clean_output_row(output_row))


CRACK_ANOMALIES_OUTPUT_COLUMNS = [
    "ID",
    "Odometer",
    "Azimuth",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Height",
    "Length",
    "Largest Individual Indication",
    "Width",
    "DepthPercent",
    "DepthPercentLD",
    "DepthPercentText",
    "Failure Pressure",
    "FPR",
    "FPRTC",
    "RPR",
    "Relative Position",
    "Type",
    "Description",
    "Surface",
    "POE",
    "Due Date",
    "Status",
    "Growth Rate Depth",
    "Growth Rate Length",
    "Discovery Date",
    "ESF",
    "Depth Tolerance - @ 80% Conf.",
    "Depth Tolerance - StdDev",
    "Length Tolerance - @ 80% Conf.",
    "Length Tolerance - StdDev",
    "Width Tolerance - @ 80% Conf.",
    "Width Tolerance - StdDev",
]


def generate_crack_anomalies_txt(ws, output_folder: str, pipe_diameter: float) -> None:
    headers = get_headers(ws)
    output_path = os.path.join(output_folder, "CrackAnomalies.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=CRACK_ANOMALIES_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()

        for row in range(2, ws.max_row + 1):
            if not is_flagged_row(ws, row, headers, ["CrackAnomalies.txt", "CrackAnomalies"]):
                continue

            maop = get_value(ws, row, headers, ["MAOP (kPa)", "MAOP (psi)"])
            mop = get_value(ws, row, headers, ["MOP (kPa)", "MOP (psi)"])
            smys = get_value(ws, row, headers, ["SMYS (kPa)", "SMYS (psi)"])
            mb31g = get_value(ws, row, headers, ["MB31G", "MB31G (kPa)", "MB31G (psi)"])
            wall_thickness = get_value(ws, row, headers, ["Wall Thickness (mm)", "Wall Thickness (in)"])

            output_row = {
                "ID": get_value(ws, row, headers, ["Feature ID", "Feature Id"]),
                "Odometer": get_value(ws, row, headers, ["Odometer Main (ft)", "Odometer Main (m)"]),
                "Azimuth": get_callbox_azimuth(ws, row, headers),
                "X_Coord": get_value(ws, row, headers, ["Easting (ft)", "Easting (m)"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (ft)", "Northing (m)"]),
                "Lat": get_value(ws, row, headers, ["Latitude", "Lat"]),
                "Long": get_value(ws, row, headers, ["Longitude", "Long"]),
                "Height": get_value(ws, row, headers, ["Height (ft)", "Height (m)"]),
                "Length": get_value(ws, row, headers, ["Length (in)", "Length (mm)", "Length"]),
                "Largest Individual Indication": get_largest_individual_indication(ws, row, headers),
                "Width": get_value(ws, row, headers, ["Width (in)", "Width (mm)", "Width"]),
                "DepthPercent": get_value(ws, row, headers, ["Depth (%)", "DepthPercent"]),
                "DepthPercentLD": "",
                "DepthPercentText": "",
                "Failure Pressure": mb31g,
                "FPR": calc_fpr_choose_pressure(mb31g, maop, mop),
                "FPRTC": get_existing_export_value(ws, row, headers, "FPRTC"),
                "RPR": calc_rpr(mb31g, smys, wall_thickness, pipe_diameter),
                "Relative Position": "",
                "Type": "",
                "Description": get_dras_description(ws, row, headers),
                "Surface": get_surface(ws, row, headers),
                "POE": "",
                "Due Date": "",
                "Status": get_existing_export_value(ws, row, headers, "Status"),
                "Growth Rate Depth": "",
                "Growth Rate Length": "",
                "Discovery Date": "",
                "ESF": "",
                "Depth Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]),
                "Depth Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]), 2),
                ),

                "Length Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]),
                "Length Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]), 2),
                ),

                "Width Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]),
                "Width Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]), 2),
                ),
            }

            writer.writerow(clean_output_row(output_row))

FACILITIES_OUTPUT_COLUMNS = [
    "ID",
    "Odometer",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Height",
    "Tool Speed",
    "Description",
]
def generate_facilities_txt(ws, output_folder: str) -> None:
    headers = get_headers(ws)
    output_path = os.path.join(output_folder, "Facilities.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=FACILITIES_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()

        for row in range(2, ws.max_row + 1):
            if not is_flagged_row(ws, row, headers, ["Facilities.txt"]):
                continue

            output_row = {
                "ID": get_value(ws, row, headers, ["Feature ID", "Feature Id"]),
                "Odometer": get_value(ws, row, headers, ["Odometer Main (ft)", "Odometer Main (m)", "Odometer"]),
                "X_Coord": get_value(ws, row, headers, ["Easting (ft)", "Easting (m)"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (ft)", "Northing (m)"]),
                "Lat": get_value(ws, row, headers, ["Latitude", "Lat"]),
                "Long": get_value(ws, row, headers, ["Longitude", "Long"]),
                "Height": get_value(ws, row, headers, ["Height (ft)", "Height (m)"]),
                "Tool Speed": get_value(ws, row, headers, ["Speed (m/s)", "Speed (ft/s)"]),
                "Description": get_facilities_description(ws, row, headers),
            }

            writer.writerow(clean_output_row(output_row))


OTHER_ANOMALIES_OUTPUT_COLUMNS = [
    "ID", "Odometer", "Azimuth", "X_Coord", "Y_Coord",
    "Lat", "Long", "Height", "Length", "Width", "Depth (Geometric)",
    "Depth (Volumetric)", "Strain", "Status", "Description", "Surface",
    "Discovery Date", "Hardness", "Depth Tolerance - @ 80% Conf.",
    "Depth Tolerance - StdDev", "Length Tolerance - @ 80% Conf.",
    "Length Tolerance - StdDev", "Width Tolerance - @ 80% Conf.",
    "Width Tolerance - StdDev",
]


def generate_other_anomalies_txt(ws, output_folder: str) -> None:
    headers = get_headers(ws)
    output_path = os.path.join(output_folder, "OtherAnomalies.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=OTHER_ANOMALIES_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )
        writer.writeheader()

        for row in range(2, ws.max_row + 1):
            if not is_flagged_row(ws, row, headers, ["OtherAnomalies.txt"]):
                continue

            feature_type = normalize_text(get_value(ws, row, headers, ["Feature Type"]))
            depth_value = get_value(ws, row, headers, ["Depth (%)"])

            output_row = {
                "ID": get_value(ws, row, headers, ["Feature ID", "Feature Id"]),
                "Odometer": get_value(ws, row, headers, ["Odometer Main (ft)", "Odometer Main (m)"]),
                "Azimuth": get_callbox_azimuth(ws, row, headers),
                "X_Coord": get_value(ws, row, headers, ["Easting (ft)", "Easting (m)"]),
                "Y_Coord": get_value(ws, row, headers, ["Northing (ft)", "Northing (m)"]),
                "Lat": get_value(ws, row, headers, ["Latitude", "Lat"]),
                "Long": get_value(ws, row, headers, ["Longitude", "Long"]),
                "Height": get_value(ws, row, headers, ["Height (ft)", "Height (m)"]),
                "Length": get_value(ws, row, headers, ["Length (mm)", "Length (in)", "Length"]),
                "Width": get_value(ws, row, headers, ["Width (mm)", "Width (in)", "Width"]),
                "Depth (Geometric)": depth_value if feature_type == "deformation" else "",
                "Depth (Volumetric)": "" if feature_type == "deformation" else depth_value,
                "Strain": get_existing_export_value(ws, row, headers, "Strain"),
                "Status": get_existing_export_value(ws, row, headers, "Status"),
                "Description": get_dras_description(ws, row, headers),
                "Surface": get_surface(ws, row, headers),
                "Discovery Date": "",
                "Hardness": "",
                "Depth Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]),
                "Depth Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Depth Tolerance - @ 80% Conf."]), 2),
                ),

                "Length Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]),
                "Length Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Length Tolerance - @ 80% Conf."]), 2,
                ),),

                "Width Tolerance - @ 80% Conf.": get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]),
                "Width Tolerance - StdDev": calc_tolerance_stddev_from_80_conf(
                    format_decimal(get_value(ws, row, headers, ["Width Tolerance - @ 80% Conf."]), 2
                ),),
            }

            writer.writerow(clean_output_row(output_row))


BEND_STRAIN_OUTPUT_COLUMNS = [
    "Feature Identifier",
    "Odometer",
    "Length",
    "Type",
    "Peak Strain Odometer",
    "Strain Orientation",
    "Total Strain",
    "Horizontal Strain",
    "Vertical Strain",
    "Horizontal Strain at Peak Strain",
    "Vertical Strain at Peak Strain",
    "Strain Direction",
    "Peak Strain Difference Odometer",
    "Strain Difference Orientation",
    "Max Total Strain Difference",
    "Max Horizontal Strain Difference",
    "Max Vertical Strain Difference",
    "Horizontal Strain Difference at Peak Strain Change",
    "Vertical Strain Difference at Peak Strain Change",
    "Strain Change Direction",
    "Max Total Pipeline Movement",
    "Max Horizontal Pipeline Movement",
    "Max Vertical Pipeline Movement",
    "X_Coord",
    "Y_Coord",
    "Lat",
    "Long",
    "Elevation",
    "Pitch",
    "Yaw",
    "Roll",
    "Positive Value Direction",
    "Comments",
]


def generate_bend_strain_txt(output_folder: str) -> None:
    """
    Placeholder until BendStrain requirements are finalized.
    Creates BendStrain.txt with headers only.
    """

    output_path = os.path.join(output_folder, "BendStrain.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=BEND_STRAIN_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )

        writer.writeheader()



INLINE_INSPECTION_OUTPUT_COLUMNS = [
    "Name",
    "Description",
    "Vendor Company",
    "Start Date Tool 1",
    "End Date Tool 1",
    "Start Date Tool 2",
    "End Date Tool 2",
    "Report Date",
    "Start Odometer",
    "End Odometer",
    "Inspection Type",
    "Call Box Interaction Rule",
    "Rupture Pressure Algorithm",
    "Coordinate Projection",
    "Coordinate Datum",
    "Height Datum",
    "Status",
    "Measured From",
    "System of Measurement",
    "SpecVersion",
]

def generate_inline_inspection_txt(output_folder: str) -> None:
    output_path = os.path.join(output_folder, "InlineInspection.txt")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=INLINE_INSPECTION_OUTPUT_COLUMNS,
            delimiter="\t",
            lineterminator="\n",
        )

        writer.writeheader()


# ---------------------------
# FACILITIES.TXT RULE HELPERS
# ---------------------------

def get_facilities_description(ws, row: int, headers: dict[str, int]) -> str:
    feature_type = safe_str(get_value(ws, row, headers, ["Feature Type"]))
    comment = safe_str(get_value(ws, row, headers, ["Comment Working", "Comments"]))
    orientation = safe_str(get_value(ws, row, headers, ["Orientation Final", "Orientation Center", "Orientation Main"]))

    return " - ".join(part for part in [feature_type, comment, orientation] if part)

# ---------------------------
# JOINT.TXT RULE HELPERS
# ---------------------------

def get_dras_seam_type(ws, row: int, headers: dict[str, int]) -> str:
    """
    DRAS Seam Type rules:

    S = if source Seam Type is spiral
    L = if not spiral and Long Seam Orientation exists
    N = if source Seam Type is seamless
    blank = if Long Seam Orientation blank and not seamless
    """

    seam_type = normalize_text(get_value(ws, row, headers, ["Seam Type"]))
    long_seam = get_value(ws, row, headers, ["Long Seam Orientation", "Long Seam Orientation (Degree)"])

    if seam_type == "spiral":
        return "S"

    if seam_type == "seamless":
        return "N"

    if long_seam not in ("", None):
        return "L"

    return ""


def get_dras_seam_position(ws, row: int, headers: dict[str, int]) -> str:
    """
    DRAS Seam Position rule:

    Use Long Seam Orientation in degrees 0-360.
    If export already has degrees, copy it.
    If export has clock format, convert to degrees.
    """

    degree_value = get_value(ws, row, headers, ["Long Seam Orientation (Degree)"])
    if degree_value not in ("", None):
        return format_integer_if_possible(degree_value)

    clock_value = get_value(ws, row, headers, ["Long Seam Orientation"])
    degrees = clock_orientation_to_degrees(clock_value)

    return "" if degrees is None else str(degrees)


def get_dras_joint_comment(ws, row: int, headers: dict[str, int]) -> str:
    """
    Comments rule:
    concatenate Feature Type and comment.
    """

    feature = safe_str(get_value(ws, row, headers, ["Feature Type"]))
    comment = safe_str(get_value(ws, row, headers, ["Comments", "Comment Working"]))

    if feature and comment:
        return f"{feature} - {comment}"

    if feature:
        return feature

    return comment


def get_mop_value(ws, row: int, headers: dict[str, int]) -> str:
    """
    MOP rule:
    use MAOP or MOP, whichever is not blank.
    """

    maop = get_value(ws, row, headers, ["MAOP (kPa)", "MAOP (psi)"])
    mop = get_value(ws, row, headers, ["MOP (kPa)", "MOP (psi)"])

    if maop not in ("", None):
        return maop

    if mop not in ("", None):
        return mop

    return ""

FT_PER_SECOND_TO_M_PER_SECOND = 0.3048

def get_detectable_length(ws, row: int, headers: dict[str, int]) -> str:
    """
    Detectable Length rule:
    If speed < 6 m/s, return 0.1.
    If speed >= 6 m/s, return blank.
    """

    speed_mps = to_float(
        get_value(ws, row, headers, ["Speed (m/s)"])
    )

    if speed_mps is None:
        speed_fps = to_float(
            get_value(ws, row, headers, ["Speed (ft/s)"])
        )

        if speed_fps is None:
            return ""

        speed_mps = speed_fps * FT_PER_SECOND_TO_M_PER_SECOND

    return "0.1" if speed_mps < 6 else ""

# ---------------------------
# Cluster.TXT RULE HELPERS
# ---------------------------

def get_dras_description(ws, row: int, headers: dict[str, int]) -> str:
    feature = safe_str(get_value(ws, row, headers, ["Feature Type"]))
    comment = safe_str(get_value(ws, row, headers, ["Comment Working", "Comments"]))

    if feature and comment:
        if comment.startswith("-"):
            return f"{feature} {comment}"
        return f"{feature} - {comment}"

    if feature:
        return feature

    return comment


def get_surface(ws, row: int, headers: dict[str, int]) -> str:
    is_external = get_value(ws, row, headers, ["Is External", "IsExternal"])

    if is_external in ("", None):
        return "U"

    text = str(is_external).strip().lower()

    if text in {"yes", "y", "true", "1"}:
        return "E"

    if text in {"no", "n", "false", "0"}:
        return "I"

    return "U"


def calc_fpr(mb31g, maop):
    mb31g_num = to_float(mb31g)
    maop_num = to_float(maop)

    if mb31g_num is None or maop_num in (None, 0):
        return ""

    return round(mb31g_num / maop_num, 3)

def calc_rpr(mb31g, smys, wall_thickness, pipe_diameter):
    mb31g_num = to_float(mb31g)
    smys_num = to_float(smys)
    wt_num = to_float(wall_thickness)
    diameter_num = to_float(pipe_diameter)

    if None in (mb31g_num, smys_num, wt_num, diameter_num):
        return ""

    if smys_num == 0 or wt_num == 0 or diameter_num == 0:
        return ""

    return round((mb31g_num / smys_num) / (2 * wt_num / diameter_num), 3)


# ---------------------------
# CALLBOX HELPERS
# ---------------------------

def get_cluster_feature_id(
    ws,
    row: int,
    headers: dict[str, int],
    id_lookup: dict[str, str]
):
    cluster = get_value(ws, row, headers, ["Cluster"])
    feature_id = get_value(ws, row, headers, ["Feature ID"])

    if cluster in ("", None):
        return feature_id

    return id_lookup.get(str(cluster).strip(), "")


def get_callbox_azimuth(ws, row: int, headers: dict[str, int]) -> str:
    degree_value = get_value(ws, row, headers, ["Orientation Center (Degree)", "Orientation Center Degree"])
    if degree_value not in ("", None):
        return format_integer_if_possible(degree_value)

    orientation_value = get_value(ws, row, headers, ["Orientation Center"])
    degrees = clock_orientation_to_degrees(orientation_value)

    return "" if degrees is None else str(degrees)


def calc_fpr_choose_pressure(mb31g, maop, mop):
    mb31g_num = to_float(mb31g)
    maop_num = to_float(maop)
    mop_num = to_float(mop)

    pressure = maop_num if maop_num not in (None, 0) else mop_num

    if mb31g_num is None or pressure in (None, 0):
        return ""

    return round(mb31g_num / pressure, 3)


def calc_erf(maop, mop, mb31g):
    pressure = to_float(maop)
    if pressure is None:
        pressure = to_float(mop)

    mb31g_num = to_float(mb31g)

    if pressure is None or mb31g_num in (None, 0):
        return ""

    return round(pressure / (0.72 * mb31g_num), 3)

# ---------------------------
# CRACK ANOMALY HELPERS
# ---------------------------
def get_largest_individual_indication(ws, row: int, headers: dict[str, int]) -> str:
    feature_type = normalize_text(get_value(ws, row, headers, ["Feature Type"]))

    if feature_type == "ncf-a" or feature_type == "ncf-b":
        return get_value(ws, row, headers, ["Width (in)", "Width (mm)", "Width"])

    if feature_type == "swf-a" or feature_type == "swf-b":
        return get_value(ws, row, headers, ["Length (in)", "Length (mm)", "Length"])

    return ""

# ---------------------------
# GENERIC HELPERS
# ---------------------------

def get_headers(ws) -> dict[str, int]:
    """
    Returns normalized header -> column number.
    """

    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            continue

        headers[normalize_header(value)] = col

    return headers


def normalize_header(value: object) -> str:
    return str(value).strip().lower()


def normalize_text(value: object) -> str:
    return safe_str(value).strip().lower()


def get_value(ws, row: int, headers: dict[str, int], possible_headers: list[str]):
    """
    Tries multiple possible header names and returns the first matching value.
    """

    for header in possible_headers:
        col = headers.get(normalize_header(header))
        if col is not None:
            value = ws.cell(row=row, column=col).value
            return "" if value is None else value

    return ""


def is_joint_row(value: object) -> bool:
    """
    Only rows where Joint = 1 are exported to Joint.txt.
    """

    if value is None:
        return False

    text = str(value).strip().lower()

    return text in {"1", "1.0", "yes", "y", "true"}


def safe_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_float(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        return float(text)
    except ValueError:
        return None


def format_integer_if_possible(value: object) -> str:
    num = to_float(value)

    if num is None:
        return safe_str(value)

    if num.is_integer():
        return str(int(num))

    return str(num)


def clock_orientation_to_degrees(value: object) -> int | None:
    """
    Converts clock orientation to degrees.

    12:00 = 0
    3:00 = 90
    6:00 = 180
    9:00 = 270
    """

    if value is None:
        return None

    text = str(value).strip()
    if not text or ":" not in text:
        return None

    try:
        hour_text, minute_text = text.split(":")
        hour = int(hour_text)
        minute = int(minute_text)

        if hour == 12:
            hour = 0

        degrees = (hour * 30) + (minute * 0.5)
        return int(round(degrees))

    except ValueError:
        return None


def clean_output_row(row: dict[str, object]) -> dict[str, str]:
    cleaned = {}

    for key, value in row.items():
        if value is None:
            cleaned[key] = ""
        else:
            text = str(value)

            # Prevent tabs/newlines from breaking tab-delimited rows
            text = (
                text.replace("\t", " ")
                    .replace("\r\n", " ")
                    .replace("\n", " ")
                    .replace("\r", " ")
                    .strip()
            )

            cleaned[key] = text

    return cleaned

def is_flagged_row(ws, row: int, headers: dict[str, int], flag_headers: list[str]) -> bool:
    value = get_value(ws, row, headers, flag_headers)
    return str(value).strip().lower() in {"1", "1.0", "yes", "y", "true"}


def get_existing_export_value(ws, row: int, headers: dict[str, int], column_name: str):
    """
    Finds and returns a value from the imported Excel/export file by exact header.
    Used for fields like FPRTC and Status where the rule says:
    'find column with this header in export file'.
    """
    return get_value(ws, row, headers, [column_name])


def calc_tolerance_stddev_from_80_conf(value):
    num = to_float(value)
    if num is None:
        return ""
    return round(num / 1.25, 2)

def format_decimal(value, places=2):
    num = to_float(value)
    if num is None:
        return ""
    return round(num, places)