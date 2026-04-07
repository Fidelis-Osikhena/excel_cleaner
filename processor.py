from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill


# ---------------------------
# HEADER HANDLING
# ---------------------------

def get_headers(ws: Worksheet) -> dict[str, int]:
    """
    Returns a dictionary mapping:
    normalized header name (lowercase, stripped) -> column index
    """
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is None:
            continue
        headers[str(value).strip().lower()] = col
    return headers


def ensure_column_after(ws: Worksheet, after_header: str, new_header: str, headers: dict[str, int]) -> None:
    """
    Inserts a column after a given header (if it doesn't already exist)
    and sets the header name.
    """
    if new_header.strip().lower() in headers:
        return

    after_col = headers.get(after_header.strip().lower())
    if after_col is None:
        return

    ws.insert_cols(after_col + 1)
    ws.cell(row=1, column=after_col + 1).value = new_header


def find_first_column(ws: Worksheet, candidates: list[str], headers: dict[str, int]) -> int | None:
    """
    Finds the first matching column from a list of possible names.
    """
    for name in candidates:
        col = headers.get(name.strip().lower())
        if col is not None:
            return col
    return None


# ---------------------------
# FORMATTING
# ---------------------------

def fill_column_green(ws: Worksheet, col: int) -> None:
    """
    Applies light green fill to a column (header + all rows).
    """
    fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in range(1, ws.max_row + 1):
        ws.cell(row=row, column=col).fill = fill


# ---------------------------
# NORMALIZATION
# ---------------------------

def normalize_feature_type(value: object) -> str:
    """
    Normalizes feature type text for consistent comparisons.
    """
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_yes(value: object) -> bool:
    """
    Returns True if value represents 'yes'.
    """
    if value is None:
        return False
    return str(value).strip().lower() in {"yes", "y", "true", "1"}


def is_girth_weld(value: object) -> bool:
    """
    Checks if feature type is 'girth weld'.
    """
    return normalize_feature_type(value) == "girth weld"


# ---------------------------
# ORIENTATION HELPERS
# ---------------------------

def clean_orientation(value: object) -> str | None:
    """
    Cleans orientation values into HH:MM format.

    Handles:
    - decimal values (e.g. 0.5 -> 12:30)
    - strings like '12:30'
    - leading 00 -> converts to 12
    """
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    # Already in HH:MM
    if ":" in text:
        try:
            h, m = text.split(":")
            h = int(h)
            m = int(m)

            if h == 0:
                h = 12

            return f"{h}:{m:02d}"
        except:
            return None

    # Decimal conversion
    try:
        val = float(text)

        total_minutes = int(round(val * 60))
        h = (total_minutes // 60) % 12
        m = total_minutes % 60

        if h == 0:
            h = 12

        return f"{h}:{m:02d}"

    except:
        return None


def orientation_to_degrees(value: str) -> int | None:
    """
    Converts HH:MM orientation to degrees (0–360).
    Returns integer (0 decimals).
    """
    if not value:
        return None

    try:
        h, m = map(int, value.split(":"))

        if h == 12:
            h = 0

        total_minutes = h * 60 + m
        degrees = (total_minutes / (12 * 60)) * 360

        return int(round(degrees))

    except:
        return None


# ---------------------------
# DEPTH
# ---------------------------

def get_depth_percent(value: object) -> float | None:
    """
    Safely parses depth percentage values.
    """
    if value is None:
        return None

    text = str(value).strip().replace("%", "")

    if not text:
        return None

    try:
        return float(text)
    except:
        return None
