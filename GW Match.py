from openpyxl import load_workbook, Workbook
from pathlib import Path

# -----------------------------
# FILE SETTINGS
# -----------------------------

input_file = Path("input.xlsx")
output_file = Path("aligned_girth_welds.xlsx")
sheet_name = "Sheet1"

# Original column layout
left_cols = ["A", "B", "C"]      # A Weld Id, A Feature Type, A Odometer Main
right_cols = ["H", "I", "J"]     # B Weld Id, B Odometer Main, B Feature Type

header_row = 1

# Column positions inside each side
left_weld_index = 0
left_feature_index = 1

right_weld_index = 0
right_feature_index = 2

girth_text = "girth weld"


# -----------------------------
# HELPER FUNCTIONS
# -----------------------------

def clean(value):
    """Clean cell text."""
    if value is None:
        return ""
    return str(value).strip()


def clean_weld_id(value):
    """Clean Weld ID values so 10 and 10.0 match."""
    value = clean(value)

    if value.endswith(".0"):
        value = value[:-2]

    return value


def is_girth(row, feature_index):
    """Check whether a row is a Girth Weld row."""
    return clean(row[feature_index]).lower() == girth_text


def blank_row(width):
    """Create a blank row with the same number of columns."""
    return [""] * width


def read_side(ws, cols):
    """Read one side of the worksheet."""
    headers = [ws[f"{col}{header_row}"].value for col in cols]
    rows = []

    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws[f"{col}{r}"].value for col in cols]

        # Keep rows that have at least one value
        if any(clean(cell) != "" for cell in row):
            rows.append(row)

    return headers, rows


def split_into_girth_blocks(rows, feature_index, weld_index):
    """
    Split rows into blocks.

    Each block starts with a Girth Weld row and includes all rows
    after it until the next Girth Weld.
    """
    pre_rows = []
    blocks = []
    current_block = None

    for row in rows:
        if is_girth(row, feature_index):
            if current_block is not None:
                blocks.append(current_block)

            current_block = {
                "weld_id": clean_weld_id(row[weld_index]),
                "rows": [row]
            }

        else:
            if current_block is None:
                pre_rows.append(row)
            else:
                current_block["rows"].append(row)

    if current_block is not None:
        blocks.append(current_block)

    return pre_rows, blocks


def align_block_sequences(left_blocks, right_blocks):
    """
    Align girth weld blocks using Weld ID.
    Matching blocks must have the same Weld ID.
    """
    n = len(left_blocks)
    m = len(right_blocks)

    # Longest Common Subsequence table
    dp = [[0] * (m + 1) for _ in range(n + 1)]

    for i in range(n - 1, -1, -1):
        for j in range(m - 1, -1, -1):
            if left_blocks[i]["weld_id"] == right_blocks[j]["weld_id"]:
                dp[i][j] = 1 + dp[i + 1][j + 1]
            else:
                dp[i][j] = max(dp[i + 1][j], dp[i][j + 1])

    operations = []
    i = 0
    j = 0

    while i < n or j < m:
        if (
            i < n
            and j < m
            and left_blocks[i]["weld_id"] == right_blocks[j]["weld_id"]
        ):
            operations.append(("match", left_blocks[i], right_blocks[j]))
            i += 1
            j += 1

        elif j < m and (i == n or dp[i][j + 1] >= dp[i + 1][j]):
            operations.append(("right_only", None, right_blocks[j]))
            j += 1

        else:
            operations.append(("left_only", left_blocks[i], None))
            i += 1

    return operations


def append_padded_rows(output_rows, left_rows, right_rows):
    """
    Add left/right rows to output, padding the shorter side with blanks.
    """
    max_len = max(len(left_rows), len(right_rows))

    for i in range(max_len):
        left_row = left_rows[i] if i < len(left_rows) else blank_row(3)
        right_row = right_rows[i] if i < len(right_rows) else blank_row(3)

        # Keep original layout:
        # A:C = left table
        # D:G = blank spacer columns
        # H:J = right table
        output_rows.append(left_row + ["", "", "", ""] + right_row)


# -----------------------------
# LOAD INPUT FILE
# -----------------------------

wb = load_workbook(input_file)
ws = wb[sheet_name]

left_headers, left_rows = read_side(ws, left_cols)
right_headers, right_rows = read_side(ws, right_cols)

left_pre, left_blocks = split_into_girth_blocks(
    left_rows,
    left_feature_index,
    left_weld_index
)

right_pre, right_blocks = split_into_girth_blocks(
    right_rows,
    right_feature_index,
    right_weld_index
)

operations = align_block_sequences(left_blocks, right_blocks)


# -----------------------------
# BUILD OUTPUT ROWS
# -----------------------------

output_rows = []

# Header row
output_rows.append(left_headers + ["", "", "", ""] + right_headers)

# Rows before first girth weld
append_padded_rows(output_rows, left_pre, right_pre)

# Girth weld blocks
for operation, left_block, right_block in operations:

    if operation == "match":
        append_padded_rows(
            output_rows,
            left_block["rows"],
            right_block["rows"]
        )

    elif operation == "left_only":
        append_padded_rows(
            output_rows,
            left_block["rows"],
            []
        )

    elif operation == "right_only":
        append_padded_rows(
            output_rows,
            [],
            right_block["rows"]
        )


# -----------------------------
# WRITE OUTPUT FILE
# -----------------------------

out_wb = Workbook()
out_ws = out_wb.active
out_ws.title = "Aligned"

for r, row in enumerate(output_rows, start=1):
    for c, value in enumerate(row, start=1):
        out_ws.cell(row=r, column=c, value=value)

out_wb.save(output_file)

print(f"Done. Saved as: {output_file}")
print(f"Left girth welds found: {len(left_blocks)}")
print(f"Right girth welds found: {len(right_blocks)}")
print(f"Matched girth welds: {sum(1 for op in operations if op[0] == 'match')}")