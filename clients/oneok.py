from __future__ import annotations

import math
from openpyxl import Workbook

from processor import (
    clean_orientation,
    ensure_column_after,
    fill_column_green,
    find_first_column,
    get_depth_percent,
    get_headers,
    is_girth_weld,
    normalize_feature_type,
    orientation_to_degrees,
    parse_yes,
)


def process_oneok_workbook(wb: Workbook, pipe_diameter: float | None = None) -> None:
    ws = wb.active

    headers = _ensure_oneok_columns(ws)
    _process_oneok_single_pass(ws, headers, pipe_diameter)


# ---------------------------
# COLUMN SETUP
# ---------------------------

def _ensure_oneok_columns(ws) -> dict[str, int]:
    headers = get_headers(ws)

    # Orientation outputs
    if "orientation final" not in headers and "orientation main" in headers:
        ensure_column_after(ws, "Orientation Main", "Orientation Final", headers)
        headers = get_headers(ws)

    if "orientation final (degree)" not in headers and "orientation final" in headers:
        ensure_column_after(ws, "Orientation Final", "Orientation Final (Degree)", headers)
        headers = get_headers(ws)

    if "long seam orientation" in headers and "long seam orientation (degree)" not in headers:
        long_col = headers["long seam orientation"]
        ws.insert_cols(long_col + 1)
        ws.cell(row=1, column=long_col + 1).value = "Long Seam Orientation (Degree)"
        headers = get_headers(ws)

    # Length/Width final
    if "length (in)" in headers and "length (in) final" not in headers:
        ensure_column_after(ws, "Length (in)", "Length (in) Final", headers)
        headers = get_headers(ws)

    if "width (in)" in headers and "width (in) final" not in headers:
        ensure_column_after(ws, "Width (in)", "Width (in) Final", headers)
        headers = get_headers(ws)

    # Depth outputs
    depth_candidates = ["Depth (%)", "Depth %", "Depth", "% Depth"]
    depth_col = find_first_column(ws, depth_candidates, headers)

    if depth_col is not None and "depth (%wt)" not in headers:
        ensure_column_after(ws, "Depth (%)", "Depth (%WT)", headers)
        headers = get_headers(ws)

    if depth_col is not None and "depth (%od)" not in headers:
        ensure_column_after(ws, "Depth (%)", "Depth (%OD)", headers)
        headers = get_headers(ws)

    # Feature / tool / comments
    if "feature type" in headers and "feature type final" not in headers:
        ensure_column_after(ws, "Feature Type", "Feature Type Final", headers)
        headers = get_headers(ws)

    if "feature type" in headers and "tool technology final" not in headers:
        ensure_column_after(ws, "Feature Type", "Tool Technology Final", headers)
        headers = get_headers(ws)

    if "comments" not in headers:
        anchor = "Feature Type Final" if "feature type final" in headers else "Feature Type"
        if anchor.lower() in headers:
            ensure_column_after(ws, anchor, "Comments", headers)
            headers = get_headers(ws)

    if "comments" in headers and "comment working 2 (gw proximity)" not in headers:
        ensure_column_after(ws, "Comments", "Comment Working 2 (GW Proximity)", headers)
        headers = get_headers(ws)

    # Additional section comment columns
    new_comment_columns = [
        "COMMENT (Casing)",
        "COMMENT (Marker Bands)",
        "COMMENT (Recoat)",
        "COMMENT (Sleeve)",
    ]

    anchor = "Comment Working 2 (GW Proximity)" if "comment working 2 (gw proximity)" in headers else "Comments"
    for col_name in new_comment_columns:
        if col_name.strip().lower() not in headers:
            ensure_column_after(ws, anchor, col_name, headers)
            headers = get_headers(ws)
        anchor = col_name

    return get_headers(ws)


# ---------------------------
# SINGLE PASS
# ---------------------------

def _process_oneok_single_pass(ws, headers: dict[str, int], pipe_diameter: float | None) -> None:
    max_row = ws.max_row

    # Source columns
    feature_col = headers.get("feature type")
    if feature_col is None:
        return

    is_marker_col = headers.get("ismarker")
    is_external_col = headers.get("is external")
    sensor_type_col = headers.get("sensor type")

    orientation_main_col = headers.get("orientation main")
    long_seam_col = headers.get("long seam orientation")
    bend_radius_col = headers.get("bend radius (xd)")
    bend_angle_col = headers.get("bend angle")
    bend_orientation_col = headers.get("bend orientation")
    length_col = headers.get("length (in)")
    width_col = headers.get("width (in)")
    depth_col = find_first_column(ws, ["Depth (%)", "Depth %", "Depth", "% Depth"], headers)

    # Output columns
    orientation_final_col = headers.get("orientation final")
    orientation_final_deg_col = headers.get("orientation final (degree)")
    long_seam_deg_col = headers.get("long seam orientation (degree)")
    length_final_col = headers.get("length (in) final")
    width_final_col = headers.get("width (in) final")
    depth_wt_col = headers.get("depth (%wt)")
    depth_od_col = headers.get("depth (%od)")
    feature_type_final_col = headers.get("feature type final")
    tool_technology_final_col = headers.get("tool technology final")
    comments_col = headers.get("comments")
    gw_proximity_col = headers.get("comment working 2 (gw proximity)")
    casing_comment_col = headers.get("comment (casing)")
    marker_bands_comment_col = headers.get("comment (marker bands)")
    recoat_comment_col = headers.get("comment (recoat)")
    sleeve_comment_col = headers.get("comment (sleeve)")

    # Delta columns
    weld_delta_main_col = _find_delta_column(headers, [
        "US Weld Δ main (ft)",
        "US Weld ? main (ft)",
        "US Weld Delta main (ft)",
    ])
    marker_delta_main_col = _find_delta_column(headers, [
        "US Marker Δ main (ft)",
        "US Marker ? main (ft)",
        "US Marker Delta main (ft)",
    ])

    # GW proximity source columns
    us_left_col = _find_delta_column(headers, [
        "US Weld Δ from left edge (ft)",
        "US Weld ? from left edge (ft)",
        "US Weld Delta from left edge (ft)",
    ])
    ds_left_col = _find_delta_column(headers, [
        "DS Weld Δ from left edge (ft)",
        "DS Weld ? from left edge (ft)",
        "DS Weld Delta from left edge (ft)",
    ])
    us_right_col = _find_delta_column(headers, [
        "US Weld Δ from right edge (ft)",
        "US Weld ? from right edge (ft)",
        "US Weld Delta from right edge (ft)",
    ])
    ds_right_col = _find_delta_column(headers, [
        "DS Weld Δ from right edge (ft)",
        "DS Weld ? from right edge (ft)",
        "DS Weld Delta from right edge (ft)",
    ])

    # Feature sets
    length_width_features = {
        "cp attachment",
        "metal object - touching",
        "pipe support - rectangular",
        "river weight",
        "stopple",
        "tap",
        "tee",
    }

    depth_wt_features = {
        "metal loss",
        "metal loss manufacturing",
        "ncf-a",
        "ncf-b",
        "sswc",
        "swa",
        "swf-a",
        "swf-b",
    }

    depth_od_features = {
        "deformation",
        "deformation - ovality",
        "deformation w/ metal loss",
    }

    gw_adjacent_features = {
        "deformation",
        "deformation - ovality",
        "deformation w/ metal loss",
        "manufacturing anomaly",
        "metal loss",
        "metal loss manufacturing",
        "ncf-a",
        "ncf-b",
        "sswc",
        "stopple",
        "swa",
    }

    section_comment_features = {
        "deformation",
        "deformation - ovality",
        "deformation w/ metal loss",
        "girth weld anomaly",
        "weld anomaly",
        "manufacturing anomaly",
        "metal loss",
        "metal loss cluster",
        "metal loss manufacturing",
        "metal loss manufacturing cluster",
        "ncf-a",
        "ncf-b",
        "sswc",
        "swa",
        "swa cluster",
        "swf-a",
        "swf-b",
    }

    # Track whether current row is inside each begin/end section
    casing_depth = 0
    marker_band_depth = 0
    recoat_depth = 0
    sleeve_depth = 0

    # One pass across rows
    for row in range(2, max_row + 1):
        feature_value = ws.cell(row=row, column=feature_col).value
        feature = normalize_feature_type(feature_value)

        is_marker_value = ws.cell(row=row, column=is_marker_col).value if is_marker_col else None
        is_external_value = ws.cell(row=row, column=is_external_col).value if is_external_col else None
        sensor_type_value = ws.cell(row=row, column=sensor_type_col).value if sensor_type_col else None
        sensor_type = normalize_feature_type(sensor_type_value)

        existing_comment = ws.cell(row=row, column=comments_col).value if comments_col else None
        existing_comment_text = "" if existing_comment is None else str(existing_comment).strip()

        source_depth_value = ws.cell(row=row, column=depth_col).value if depth_col else None

        # -------------------
        # Orientation block
        # -------------------
        orientation_final_val = ""
        orientation_final_deg_val = ""
        long_seam_cleaned_val = ""
        long_seam_deg_val = ""

        # Build Orientation Final from Orientation Main first
        if orientation_main_col:
            main_value = ws.cell(row=row, column=orientation_main_col).value
            cleaned_main = clean_orientation(main_value)

            if cleaned_main:
                orientation_final_val = cleaned_main
                deg = orientation_to_degrees(cleaned_main)
                orientation_final_deg_val = "" if deg is None else deg

        # Process Long Seam Orientation, preserving non-girth rows
        if long_seam_col and long_seam_deg_col:
            long_value = ws.cell(row=row, column=long_seam_col).value

            # Fill down from the row above if blank
            if long_value is None or str(long_value).strip() == "":
                if row > 2:
                    long_value = ws.cell(row=row - 1, column=long_seam_col).value

            cleaned_long = clean_orientation(long_value)

            if cleaned_long:
                long_seam_cleaned_val = cleaned_long
                deg = orientation_to_degrees(cleaned_long)
                long_seam_deg_val = "" if deg is None else deg

                # Write cleaned/fill-down value back into Long Seam Orientation
                ws.cell(row=row, column=long_seam_col).value = cleaned_long

                # ONLY override Orientation Final for Girth Weld
                if is_girth_weld(feature_value):
                    orientation_final_val = cleaned_long
                    orientation_final_deg_val = long_seam_deg_val
            else:
                long_seam_cleaned_val = ""
                long_seam_deg_val = ""

        if orientation_final_col:
            ws.cell(row=row, column=orientation_final_col).value = orientation_final_val
            ws.cell(row=row, column=orientation_final_col).number_format = "@"

        if orientation_final_deg_col:
            ws.cell(row=row, column=orientation_final_deg_col).value = orientation_final_deg_val
            ws.cell(row=row, column=orientation_final_deg_col).number_format = "0"

        if long_seam_col:
            ws.cell(row=row, column=long_seam_col).value = long_seam_cleaned_val
            ws.cell(row=row, column=long_seam_col).number_format = "@"

        if long_seam_deg_col:
            ws.cell(row=row, column=long_seam_deg_col).value = long_seam_deg_val
            ws.cell(row=row, column=long_seam_deg_col).number_format = "0"

        # -------------------
        # Delta block
        # -------------------
        if weld_delta_main_col and is_girth_weld(feature_value):
            ws.cell(row=row, column=weld_delta_main_col).value = 0

        if marker_delta_main_col and parse_yes(is_marker_value):
            ws.cell(row=row, column=marker_delta_main_col).value = 0

        # -------------------
        # Length / Width Final
        # -------------------
        if length_final_col:
            length_final = ""
            if feature in length_width_features and length_col:
                length_final = _round_to_zero_decimal(ws.cell(row=row, column=length_col).value)
                length_final = "" if length_final is None else length_final
            ws.cell(row=row, column=length_final_col).value = length_final
            ws.cell(row=row, column=length_final_col).number_format = "0"

        if width_final_col:
            width_final = ""
            if feature in length_width_features and width_col:
                width_final = _round_to_zero_decimal(ws.cell(row=row, column=width_col).value)
                width_final = "" if width_final is None else width_final
            ws.cell(row=row, column=width_final_col).value = width_final
            ws.cell(row=row, column=width_final_col).number_format = "0"

        # -------------------
        # Depth outputs
        # -------------------
        if depth_wt_col:
            depth_wt_value = source_depth_value if feature in depth_wt_features and source_depth_value is not None else ""
            ws.cell(row=row, column=depth_wt_col).value = depth_wt_value

        if depth_od_col:
            depth_od_value = source_depth_value if feature in depth_od_features and source_depth_value is not None else ""
            ws.cell(row=row, column=depth_od_col).value = depth_od_value

        # -------------------
        # Comments
        # -------------------
        comment = ""
        gw_comment = ""
        casing_comment = ""
        marker_bands_comment = ""
        recoat_comment = ""
        sleeve_comment = ""

        if feature in {
            "repair marker begin",
            "agm",
            "casing begin",
            "girth weld",
            "marker band begin",
            "recoat begin",
            "sleeve begin",
        }:
            comment = existing_comment_text

        elif feature == "valve":
            comment = existing_comment_text if parse_yes(is_marker_value) else (str(feature_value).strip() if feature_value else "")

        elif feature == "bend":
            radius_val = ws.cell(row=row, column=bend_radius_col).value if bend_radius_col else None
            angle_val = ws.cell(row=row, column=bend_angle_col).value if bend_angle_col else None
            orientation_val = ws.cell(row=row, column=bend_orientation_col).value if bend_orientation_col else None
            comment = comment_bend(radius_val, angle_val, orientation_val)

        elif feature in {
            "half sole repair",
            "magnet",
            "patch repair",
            "pipe support - rectangular",
            "puddle weld repair",
            "repair marker end",
            "river weight",
            "casing end",
            "flange",
            "marker band end",
            "pipe support - circumferential",
            "recoat end",
            "sleeve end",
        }:
            comment = str(feature_value).strip() if feature_value else ""

        elif feature in {"deformation", "deformation - ovality", "deformation w/ metal loss"}:
            depth_percent = get_depth_percent(source_depth_value)

            if depth_percent is not None and pipe_diameter is not None:
                depth_inches = (depth_percent / 100.0) * pipe_diameter
                depth_inches_text = f"{depth_inches:.2f} inch"

                if depth_percent >= 1.0:
                    if feature == "deformation w/ metal loss":
                        comment = f"Dent - {depth_inches_text}, w/ Possible Associated Metal Loss"
                    else:
                        comment = f"Dent - {depth_inches_text}"
                else:
                    comment = f"Geometric Anomaly - {depth_inches_text}"
            else:
                comment = "Deformation"

        elif feature == "manufacturing anomaly":
            comment = existing_comment_text

        elif feature in {"metal loss", "metal loss manufacturing", "swa"}:
            side = "External" if parse_yes(is_external_value) else "Internal"
            comment = f"Metal Loss - {side}"

        elif feature in {"ncf-a", "ncf-b"}:
            length_val = ws.cell(row=row, column=length_col).value if length_col else None
            width_val = ws.cell(row=row, column=width_col).value if width_col else None
            comment = comment_ncf(length_val, width_val)

        elif feature == "sswc":
            comment = "Possible selective seam weld corrosion"

        elif feature == "weld anomaly":
            comment = "Possible Girth Weld Anomaly"

        elif feature == "stopple":
            comment = comment_stopple(orientation_final_val)

        elif feature == "tap":
            comment = comment_tap(orientation_final_val)

        elif feature == "tee":
            comment = comment_tee(orientation_final_val)

        # -------------------
        # Feature Type Final
        # -------------------
        feature_type_final_val = map_feature_type_final(
            feature_value=feature_value,
            is_marker_value=is_marker_value,
            depth_value=source_depth_value,
        )

        # Override: Metal Object Touching + Attachment
        if feature == "metal object - touching" and "attachment" in existing_comment_text.lower():
            feature_type_final_val = "LOCATION"

        if feature_type_final_col:
            ws.cell(row=row, column=feature_type_final_col).value = feature_type_final_val

        # -------------------
        # Tool Technology Final
        # -------------------
        tool_technology_final_val = map_tool_technology_final(feature, sensor_type)

        if tool_technology_final_col:
            ws.cell(row=row, column=tool_technology_final_col).value = tool_technology_final_val

        # -------------------
        # GW proximity column
        # -------------------
        if feature in gw_adjacent_features:
            if is_adjacent_to_girth_weld(
                ws, row,
                us_left_col,
                ds_left_col,
                us_right_col,
                ds_right_col,
            ):
                gw_comment = "adjacent to girth weld"

        # -------------------
        # New section comment columns
        # -------------------
        if feature in section_comment_features:
            if casing_depth > 0:
                casing_comment = "feature appears in a Casing"
            if marker_band_depth > 0:
                marker_bands_comment = "between marker bands"
            if recoat_depth > 0:
                recoat_comment = "feature appears in area of recoat"
            if sleeve_depth > 0:
                sleeve_comment = "under sleeve"

        if comments_col:
            ws.cell(row=row, column=comments_col).value = comment

        if gw_proximity_col:
            ws.cell(row=row, column=gw_proximity_col).value = gw_comment

        if casing_comment_col:
            ws.cell(row=row, column=casing_comment_col).value = casing_comment

        if marker_bands_comment_col:
            ws.cell(row=row, column=marker_bands_comment_col).value = marker_bands_comment

        if recoat_comment_col:
            ws.cell(row=row, column=recoat_comment_col).value = recoat_comment

        if sleeve_comment_col:
            ws.cell(row=row, column=sleeve_comment_col).value = sleeve_comment

        # Update begin/end state AFTER processing the row
        if feature == "casing begin":
            casing_depth += 1
        elif feature == "casing end" and casing_depth > 0:
            casing_depth -= 1

        if feature == "marker band begin":
            marker_band_depth += 1
        elif feature == "marker band end" and marker_band_depth > 0:
            marker_band_depth -= 1

        if feature == "recoat begin":
            recoat_depth += 1
        elif feature == "recoat end" and recoat_depth > 0:
            recoat_depth -= 1

        if feature == "sleeve begin":
            sleeve_depth += 1
        elif feature == "sleeve end" and sleeve_depth > 0:
            sleeve_depth -= 1

    # Highlight output columns once at the end
    for name in [
        "orientation final",
        "orientation final (degree)",
        "long seam orientation",
        "long seam orientation (degree)",
        "length (in) final",
        "width (in) final",
        "depth (%wt)",
        "depth (%od)",
        "feature type final",
        "tool technology final",
        "comments",
        "comment working 2 (gw proximity)",
        "comment (casing)",
        "comment (marker bands)",
        "comment (recoat)",
        "comment (sleeve)",
    ]:
        col = headers.get(name)
        if col:
            fill_column_green(ws, col)

    for col in [weld_delta_main_col, marker_delta_main_col]:
        if col:
            fill_column_green(ws, col)


# ---------------------------
# CLASSIFICATION HELPERS
# ---------------------------

def map_feature_type_final(*, feature_value: object, is_marker_value: object, depth_value: object) -> str:
    feature = normalize_feature_type(feature_value)

    if feature == "bend":
        return "LOCATION"

    if feature == "cp attachment":
        return "GAIN"

    if feature == "crack-like":
        return ""

    if feature in {"deformation", "deformation - ovality", "deformation w/ metal loss"}:
        depth_percent = get_depth_percent(depth_value)
        if depth_percent is not None and depth_percent >= 1.0:
            return "DENT"
        return "GMA"

    if feature in {
        "half sole repair",
        "magnet",
        "ncf-a",
        "ncf-b",
        "patch repair",
        "puddle weld repair",
        "repair marker begin",
        "repair marker end",
        "river weight",
        "weld anomaly",
        "recoat begin",
        "recoat end",
    }:
        return "MISC"

    if feature == "manufacturing anomaly":
        return "MILL ANOMALY"

    if feature in {"metal loss", "metal loss manufacturing"}:
        return "GROUP"

    if feature in {"metal object - close", "metal object - touching"}:
        return "GAIN"

    if feature in {
        "pipe support - rectangular",
        "stopple",
        "tap",
        "tee",
        "casing begin",
        "casing end",
        "flange",
        "marker band begin",
        "marker band end",
        "pipe support - circumferential",
        "sleeve begin",
        "sleeve end",
        "valve",
    }:
        if feature == "valve":
            return "AGM" if parse_yes(is_marker_value) else "LOCATION"
        return "LOCATION"

    if feature in {"sswc", "swa"}:
        return "SW GROUP"

    if feature == "swf-a":
        return "SW-A"

    if feature == "swf-b":
        return "SW-B"

    if feature == "agm":
        return "AGM"

    if feature == "girth weld":
        return "WELD"

    if feature == "coupling":
        return ""

    return ""


def map_tool_technology_final(feature: str, sensor: str) -> str:
    if feature in {"sswc", "swf-a", "swf-b"}:
        return "CMFL"

    if feature in {"deformation", "deformation - ovality", "deformation w/ metal loss"}:
        return "Geometry"

    if feature in {"ncf-a", "ncf-b"}:
        return "IDD-SM"

    if feature in {"swa", "weld anomaly"}:
        if sensor == "axial":
            return "AMFL"
        if sensor == "circumferential":
            return "CMFL"

    if feature == "manufacturing anomaly":
        if sensor == "axial":
            return "AMFL"
        if sensor == "circumferential":
            return "CMFL"
        if sensor == "geometry":
            return "Geometry"

    return ""


# ---------------------------
# COMMENT HELPERS
# ---------------------------

def get_depth_percent(value: object) -> float | None:
    if value is None:
        return None

    text = str(value).strip().replace("%", "")
    if not text:
        return None

    try:
        num = float(text)
    except (TypeError, ValueError):
        return None

    # Excel percent-style values often come through as decimals
    # Example: 0.004 means 0.4%
    if 0 < num < 1:
        return num * 100

    return num


def comment_bend(radius_value: object, angle_value: object, orientation_value: object) -> str:
    """
    Bend comment format:
    Bend - {radius 0.0}D - {angle 0.0}° - {up/right/down/left}
    """
    radius = format_one_decimal(radius_value)
    angle = format_one_decimal(angle_value)
    direction_text = bend_orientation_to_direction(orientation_value)

    # Add D and degree symbol
    radius_text = f"{radius}D" if radius else ""
    angle_text = f"{angle}°" if angle else ""
    
    return f"Bend - {radius_text} - {angle_text} - {direction_text}"


def comment_ncf(length_value: object, width_value: object) -> str:
    degrees = ncf_degrees_from_y_axis(length_value, width_value)
    if degrees is None:
        return "Possible Narrow Circumferential Feature"
    return f"Possible Narrow Circumferential Feature, {degrees} degrees from y-axis"


def comment_stopple(orientation_value: object) -> str:
    position = orientation_position(orientation_value)
    if position == "top":
        return "STOPPLE on top of pipe"
    if position == "90":
        return "STOPPLE at 90 degrees"
    if position == "bottom":
        return "STOPPLE on bottom of pipe"
    if position == "270":
        return "STOPPLE at 270 degrees"
    return "STOPPLE"


def comment_tap(orientation_value: object) -> str:
    position = orientation_position(orientation_value)
    if position == "top":
        return "Fitting on top of pipe"
    if position == "90":
        return "Tap at 90 degrees"
    if position == "bottom":
        return "Fitting on bottom of pipe"
    if position == "270":
        return "Tap at 270 degrees"
    return "Tap"


def comment_tee(orientation_value: object) -> str:
    position = orientation_position(orientation_value)
    if position == "top":
        return "TEE on top of pipe"
    if position == "90":
        return "TEE at 90 degrees"
    if position == "bottom":
        return "TEE on bottom of pipe"
    if position == "270":
        return "TEE at 270 degrees"
    return "TEE"


# ---------------------------
# GENERIC HELPERS
# ---------------------------

def format_one_decimal(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return f"{float(text):.1f}"
    except (ValueError, TypeError):
        return text


def bend_orientation_to_direction(value: object) -> str:
    cleaned = clean_orientation(value)
    if not cleaned:
        return ""

    hour, minute = map(int, cleaned.split(":"))
    total_minutes = _clock_minutes(hour, minute)

    if total_minutes >= 630 or total_minutes <= 89:
        return "up"
    if 90 <= total_minutes <= 269:
        return "right"
    if 270 <= total_minutes <= 449:
        return "down"
    if 450 <= total_minutes <= 629:
        return "left"
    return ""


def orientation_position(value: object) -> str:
    cleaned = clean_orientation(value)
    if not cleaned:
        return ""

    hour, minute = map(int, cleaned.split(":"))
    total_minutes = _clock_minutes(hour, minute)

    if total_minutes >= 630 or total_minutes <= 89:
        return "top"
    if 90 <= total_minutes <= 269:
        return "90"
    if 270 <= total_minutes <= 449:
        return "bottom"
    if 450 <= total_minutes <= 629:
        return "270"
    return ""


def _clock_minutes(hour: int, minute: int) -> int:
    if hour == 12:
        hour = 0
    return hour * 60 + minute


def ncf_degrees_from_y_axis(length_value: object, width_value: object) -> str | None:
    try:
        length_num = float(length_value)
        width_num = float(width_value)
    except (TypeError, ValueError):
        return None

    if width_num == 0:
        return None

    degrees = math.degrees(math.atan(length_num / width_num))
    return f"{degrees:.0f}"


def is_adjacent_to_girth_weld(ws, row: int, us_left_col, ds_left_col, us_right_col, ds_right_col) -> bool:
    for col in [us_left_col, ds_left_col, us_right_col, ds_right_col]:
        if not col:
            continue

        value = ws.cell(row=row, column=col).value
        num = _to_float(value)

        if num is not None and abs(num) <= 0.04:
            return True

    return False


def _round_to_zero_decimal(value: object) -> int | None:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    try:
        return int(round(float(text)))
    except (TypeError, ValueError):
        return None


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _find_delta_column(headers: dict[str, int], candidates: list[str]) -> int | None:
    for candidate in candidates:
        col = headers.get(candidate.strip().lower())
        if col is not None:
            return col
    return None
